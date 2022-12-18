from bs4 import BeautifulSoup
import requests

#load data to excel file
import openpyxl

excel = openpyxl.Workbook()
excel.sheetnames
sheet = excel.active
sheet.title = 'Top Rated Movies'
excel.sheetnames

#create column names

sheet.append(['Rank','Name','Year','IMDB Rating'])

#request module to access this website
#source will have html source code of website
try:
    source = requests.get('https://www.imdb.com/chart/top/')  #going to access the url and return a response object and this 
    #response object will be saved in source variable # the response object will the html source code of the webiste
    
    source.raise_for_status()
    
    soup = BeautifulSoup(source.text,'html.parser')
    #print(soup)
    
    movies = soup.find('tbody', class_ = "lister-list").find_all('tr')
    #print(len(movies))
    
    for movie in movies:
        
        name = movie.find('td',class_="titleColumn").a.text
        rank = movie.find('td',class_ = "titleColumn").get_text(strip=True).split('.')[0]
        year = movie.find('td', class_ = "titleColumn").span.text.strip('()')
        rating = movie.find('td', class_ = "ratingColumn").strong.text
        
        print(rank,name,year, rating)
        sheet.append([rank,name,year,rating])
        
            
except Exception as e:
    print(e)
    
excel.save('IMDB Ratings.xlsx')